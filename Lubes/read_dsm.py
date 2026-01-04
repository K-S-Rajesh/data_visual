import openpyxl
import pandas as pd

#
excel_file=r'C:\Users\ksraj\OneDrive - Indian Oil Corporation Limited\Business Plan PPT\July 24\Lubes.xlsx'
df=pd.read_excel(excel_file,'Lubes DSM')
print(df.columns)

# wb=openpyxl.load_workbook(r'C:\Users\ksraj\OneDrive - Indian Oil Corporation Limited\Business Plan PPT\July 24\Lubes.xlsx')
# # print(wb.sheetnames)
# ws=wb['Lubes DSM']
# print(ws)
# value_range=ws['A1':'B1']
# # print(value_range)
#
# for a,b in value_range:
#     print(a.value,b.value)

