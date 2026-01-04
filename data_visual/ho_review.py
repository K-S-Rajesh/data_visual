from openpyxl import Workbook,load_workbook

wb=load_workbook(r'C:\Users\ksraj\Downloads\ACTION PLAN KESO.xlsx')
sheet3=wb.sheetnames[3]
wb.active=wb[sheet3]
ws=wb.active
# print(ws['B1'].value)
print(list(ws.values))

# print(wb.sheetnames[0])



# for sheet in wb.worksheets:
#     print(sheet)


# print(wb.sheetnames())
