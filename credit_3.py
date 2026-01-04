import os
import tkinter as tk
import time
import sys
import pyautogui as pag
import pandas as pd
import numpy as np
import openpyxl


#WIDTH = 200

#HEIGHT = 300

#root = tk.Tk()

#canvas = tk.Canvas(root, height=HEIGHT, width=WIDTH)

#canvas.pack()x

#def myClick1():

#excel_file='C:\\Users\\ksraj\\Downloads\\fd32.xlsx'

#df=pd.read_excel(excel_file)
wb = openpyxl.load_workbook("C:\\Users\\ksraj\\Downloads\\fd32.xlsx")
sh = wb.active


#sys.exit()
#a=df.iat[0,0]

#print(a)
#sys.exit()

#for i in range(6):
    #a=(df.iat[i,0])
    #time.sleep(1)
    #print(a)
 #   arr=df.to_numpy()
  #  pag.write((arr[i][0]))





#sys.exit()



s.system('start sapshcut.exe -system=PRD -client=310 -user=00500230 -pw=Iocl@072023')

i=1
while i<=1:
    #im = pag.screenshot("Full screen.png")
    try:
        x, y = pag.locateCenterOnScreen("saplogonentry.png", confidence=0.7)
    except TypeError:
        i = 1
    else:
        break


time.sleep(1)
pag.hotkey('alt','space')
time.sleep(0.3)
pag.press('x')
time.sleep(1)
#sys.exit()


review_date=pag.prompt(text='Enter Review date', title='MSG Box' , default='')
time.sleep(1)
monitoring_date=pag.prompt(text='Enter Monitoring date', title='MSG Box' , default='')
time.sleep(1)
pag.hotkey('ctrl', '/')
time.sleep(1)
pag.write('/nfd32')
time.sleep(1)
pag.press('enter')
time.sleep(2)
c3 = sh.cell(row=2, column=1)
time.sleep(2)
a = str(c3.value)
time.sleep(1)
pag.write(a)
#sys.exit()
pag.press('tab')
time.sleep(1)
pag.write('c004')
time.sleep(1)
pag.press('tab', presses=4)
time.sleep(1)
pag.press('space')
time.sleep(1)
pag.press('enter')
time.sleep(1)
pag.hotkey('ctrl','tab')
time.sleep(1)
pag.press('tab', presses=5)
time.sleep(1)
pag.write(review_date)
time.sleep(1)
pag.press('tab', presses=5)
time.sleep(1)
pag.write(monitoring_date)
time.sleep(1)
pag.hotkey('ctrl','s')
time.sleep(1)
pag.press('enter')
time.sleep(1)
pag.press('enter')


#sys.exit()





for i in range(3):
    c4 = sh.cell(row=i+2,column=1)
    time.sleep(1)
    b=str(c4.value)
    time.sleep(1)
    pag.hotkey('ctrl','/')
    time.sleep(1)
    pag.write('/nfd32')
    time.sleep(1)
    pag.press('enter')
    time.sleep(1)
    pag.write(b)
    time.sleep(1)
    pag.press('tab')
    time.sleep(1)
    pag.write('c004')
    time.sleep(1)
    pag.press('tab',presses=4)
    time.sleep(1)
    #pag.press('space')
    time.sleep(1)
    pag.press('enter')
    time.sleep(1)
    pag.hotkey('ctrl','tab')
    time.sleep(1)
    pag.press('tab', presses=5)
    time.sleep(1)
    pag.write(review_date)
    time.sleep(1)
    pag.press('tab', presses=5)
    time.sleep(1)
    pag.write(monitoring_date)
    time.sleep(1)
    pag.hotkey('ctrl','s')
    time.sleep(1)
    pag.press('enter')
    time.sleep(1)
    pag.press('enter')




